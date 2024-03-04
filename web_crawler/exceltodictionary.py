#körs om man ska göra excelfil till dictionary
import openpyxl
import json 

loadpynace = openpyxl.load_workbook("latestnacecode.xlsx")

callnace = loadpynace.active

    
nacedata = {}
#print(str(callnace.max_row), str(callnace.max_column)) #radder och columber i excelfilen
nacenummer = [callnace.cell(row=i, column=1).value for i in range(1, callnace.max_row)]
nacenamn = [callnace.cell(row=i, column=2).value for i in range(1, callnace.max_row)]

for key in nacenamn:
    for value in nacenummer:
        nacedata[key] = value
        nacenummer.remove(value)
        break

nacedata.popitem()
#print (nacedata)
def valuesfordropdown():
    return nacedata
  
